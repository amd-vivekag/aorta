"""
HW Queue Eval Excel report generator.

Generates Excel reports for:
- Single run analysis (Mode A)
- Sweep analysis (Mode B)
- Multi-workload comparison (Mode C)
"""

from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import data classes from loader
from ..processing.hwqueue_loader import SingleRunData, SweepData


# Color constants
RED = "F8696B"
WHITE = "FFFFFF"
GREEN = "63BE7B"
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _sanitize_table_name(name: str) -> str:
    """Create valid Excel table name."""
    table_name = name.replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
    if not table_name[0].isalpha():
        table_name = "Tbl_" + table_name
    return table_name[:255]


def _add_excel_table(worksheet, table_name: str, start_row: int = 1) -> bool:
    """Convert worksheet data to Excel table format."""
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    if max_row <= start_row:
        return False

    # Ensure all column headers are strings
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=start_row, column=col_idx)
        if cell.value is not None and not isinstance(cell.value, str):
            cell.value = str(cell.value)

    start_cell = f"A{start_row}"
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    table_ref = f"{start_cell}:{end_cell}"

    try:
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        return True
    except Exception as e:
        print(f"    Warning: Could not create table {table_name}: {e}")
        return False


def _apply_formatting(output_path: Path, verbose: bool = False) -> None:
    """Apply formatting to the Excel file after writing."""
    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.max_row <= 1:
            continue

        # Create unique table name
        table_name = _sanitize_table_name(f"HWQ_{sheet_name}")
        if _add_excel_table(ws, table_name):
            if verbose:
                print(f"    Converted to table: {sheet_name}")

        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in range(1, ws.max_row + 1):
                try:
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(output_path)


def generate_single_run_excel(
    data: SingleRunData,
    output_path: Path,
    verbose: bool = False,
) -> Path:
    """
    Generate Excel report for single run analysis (Mode A).

    Args:
        data: SingleRunData object from loader
        output_path: Path to output Excel file

    Returns:
        Path to generated Excel file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if verbose:
        print(f"  Generating single run Excel: {output_path.name}")

    # Build summary data
    summary_data = {
        "Metric": [
            "Workload",
            "Stream Count",
            "Throughput",
            "Total Time (ms)",
            "Latency Mean (ms)",
            "Latency P50 (ms)",
            "Latency P95 (ms)",
            "Latency P99 (ms)",
            "Latency Min (ms)",
            "Latency Max (ms)",
            "Latency Std (ms)",
        ],
        "Value": [
            data.workload_name,
            data.stream_count,
            f"{data.throughput:.2f} {data.throughput_unit}",
            f"{data.total_time_ms:.3f}",
            f"{data.latency.mean:.3f}",
            f"{data.latency.p50:.3f}",
            f"{data.latency.p95:.3f}",
            f"{data.latency.p99:.3f}",
            f"{data.latency.min:.3f}",
            f"{data.latency.max:.3f}",
            f"{data.latency.std:.3f}",
        ],
    }

    # Add switch latency if available
    if data.switch_latency:
        summary_data["Metric"].extend([
            "Inter-Stream Gap (ms)",
            "Intra-Stream Gap (ms)",
            "Switch Overhead (ms)",
        ])
        summary_data["Value"].extend([
            f"{data.switch_latency.inter_stream_gap_ms:.3f}",
            f"{data.switch_latency.intra_stream_gap_ms:.3f}",
            f"{data.switch_latency.estimated_switch_overhead_ms:.3f}",
        ])

    # Add memory if available
    if data.memory:
        summary_data["Metric"].extend([
            "Peak Allocated (GB)",
            "Peak Reserved (GB)",
        ])
        summary_data["Value"].extend([
            f"{data.memory.peak_allocated_gb:.2f}",
            f"{data.memory.peak_reserved_gb:.2f}",
        ])

    summary_df = pd.DataFrame(summary_data)

    # Build per-stream data if available
    sheets = {"Summary": summary_df}

    if data.per_stream_times_ms:
        stream_data = {
            "Stream": list(range(len(data.per_stream_times_ms))),
            "Time (ms)": data.per_stream_times_ms,
        }
        sheets["Per_Stream_Times"] = pd.DataFrame(stream_data)

    # Write to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply formatting
    _apply_formatting(output_path, verbose)

    if verbose:
        print(f"    Created {len(sheets)} sheet(s)")

    return output_path


def generate_sweep_excel(
    data: SweepData,
    output_path: Path,
    verbose: bool = False,
) -> Path:
    """
    Generate Excel report for sweep analysis (Mode B).

    Args:
        data: SweepData object from loader
        output_path: Path to output Excel file

    Returns:
        Path to generated Excel file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if verbose:
        print(f"  Generating sweep Excel: {output_path.name}")

    # Build summary sheet
    best_streams, best_throughput = data.get_best_throughput()

    summary_data = {
        "Metric": [
            "Workload",
            "Stream Configurations Tested",
            "Best Stream Count",
            "Peak Throughput",
            "Inflection Point",
        ],
        "Value": [
            data.workload_name,
            len(data.results),
            best_streams,
            f"{best_throughput:.2f}",
            data.analysis.inflection_point if data.analysis.inflection_point else "N/A",
        ],
    }

    # Add environment info if available
    if data.environment.hostname:
        summary_data["Metric"].extend([
            "Hostname",
            "GPU Count",
            "HIP Version",
            "PyTorch Version",
        ])
        summary_data["Value"].extend([
            data.environment.hostname,
            data.environment.gpu_count,
            data.environment.hip_version or "N/A",
            data.environment.torch_version or "N/A",
        ])

    summary_df = pd.DataFrame(summary_data)

    # Build scaling data sheet
    scaling_rows = []
    for result in data.results:
        row = {
            "Stream_Count": result.stream_count,
            "Throughput": result.throughput,
            "Throughput_Unit": result.throughput_unit,
            "Latency_Mean_ms": result.latency.mean,
            "Latency_P50_ms": result.latency.p50,
            "Latency_P95_ms": result.latency.p95,
            "Latency_P99_ms": result.latency.p99,
            "Total_Time_ms": result.total_time_ms,
        }

        # Add efficiency if available from analysis
        if data.analysis.stream_counts and result.stream_count in data.analysis.stream_counts:
            idx = data.analysis.stream_counts.index(result.stream_count)
            if idx < len(data.analysis.efficiencies):
                row["Efficiency"] = data.analysis.efficiencies[idx]

        # Add switch latency if available
        if result.switch_latency:
            row["Inter_Stream_Gap_ms"] = result.switch_latency.inter_stream_gap_ms
            row["Intra_Stream_Gap_ms"] = result.switch_latency.intra_stream_gap_ms
            row["Switch_Overhead_ms"] = result.switch_latency.estimated_switch_overhead_ms

        scaling_rows.append(row)

    scaling_df = pd.DataFrame(scaling_rows)

    # Sort by stream count
    if not scaling_df.empty:
        scaling_df = scaling_df.sort_values("Stream_Count").reset_index(drop=True)

    sheets = {
        "Summary": summary_df,
        "Scaling_Data": scaling_df,
    }

    # Write to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply formatting
    _apply_formatting(output_path, verbose)

    if verbose:
        print(f"    Created {len(sheets)} sheet(s)")
        print(f"    Scaling data: {len(scaling_rows)} configurations")

    return output_path


def generate_hwqueue_excel(
    data: SingleRunData | SweepData,
    output_path: Path,
    verbose: bool = False,
) -> Path:
    """
    Generate Excel report for single run or sweep data.

    Automatically detects data type and calls the appropriate generator.

    Args:
        data: SingleRunData or SweepData object
        output_path: Path to output Excel file

    Returns:
        Path to generated Excel file
    """
    if isinstance(data, SweepData):
        return generate_sweep_excel(data, output_path, verbose)
    elif isinstance(data, SingleRunData):
        return generate_single_run_excel(data, output_path, verbose)
    else:
        raise ValueError(f"Unknown data type: {type(data)}")


# =============================================================================
# Comparison Mode (Mode C) - Multi-Workload Comparison
# =============================================================================


def _apply_comparison_formatting(
    output_path: Path,
    change_columns: List[Tuple[str, str]],  # (sheet_name, column_name)
    verbose: bool = False,
) -> None:
    """Apply formatting to comparison Excel file with color-coded change columns."""
    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.max_row <= 1:
            continue

        # Create unique table name
        table_name = _sanitize_table_name(f"HWQ_{sheet_name}")
        if _add_excel_table(ws, table_name):
            if verbose:
                print(f"    Converted to table: {sheet_name}")

        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in range(1, ws.max_row + 1):
                try:
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except Exception as e:
                    if verbose:
                        print(
                            f"    Warning: Could not read cell {sheet_name}!"
                            f"{col_letter}{row} for width calculation: {e}"
                        )
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Apply color scale to change columns
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header and ("Change" in str(header) or "Δ" in str(header)):
                col_letter = get_column_letter(col_idx)
                data_range = f"{col_letter}2:{col_letter}{ws.max_row}"

                try:
                    ws.conditional_formatting.add(
                        data_range,
                        ColorScaleRule(
                            start_type="min",
                            start_color=RED,
                            mid_type="num",
                            mid_value=0,
                            mid_color=WHITE,
                            end_type="max",
                            end_color=GREEN,
                        ),
                    )
                    if verbose:
                        print(f"    Applied color scale to {sheet_name}.{header}")
                except Exception as e:
                    if verbose:
                        print(f"    Warning: Could not apply formatting to {header}: {e}")

    # Move Summary to first position if it exists
    if "Summary" in wb.sheetnames:
        summary_sheet = wb["Summary"]
        wb.move_sheet(summary_sheet, offset=-(len(wb.sheetnames) - 1))
        wb.active = 0

    wb.save(output_path)


def _compute_comparison_metrics(
    baseline_data: Dict[str, SweepData],
    test_data: Dict[str, SweepData],
    common_workloads: List[str],
    threshold: float,
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    Compute comparison metrics for all common workloads.

    Returns:
        (summary_rows, regressions, improvements)
    """
    summary_rows = []
    regressions = []
    improvements = []

    for workload in common_workloads:
        baseline = baseline_data[workload]
        test = test_data[workload]

        # Get best throughput for each
        b_best_streams, b_best_throughput = baseline.get_best_throughput()
        t_best_streams, t_best_throughput = test.get_best_throughput()

        # Calculate change
        if b_best_throughput > 0:
            throughput_change = (t_best_throughput - b_best_throughput) / b_best_throughput
        else:
            throughput_change = 0

        # Determine status
        if throughput_change < -threshold:
            status = "⚠ REGRESSION"
        elif throughput_change > threshold:
            status = "✓ IMPROVED"
        else:
            status = "✓ OK"

        summary_rows.append({
            "Workload": workload,
            "Best_Streams_Base": b_best_streams,
            "Best_Streams_Test": t_best_streams,
            "Throughput_Base": round(b_best_throughput, 2),
            "Throughput_Test": round(t_best_throughput, 2),
            "Change_%": round(throughput_change * 100, 2),
            "Status": status,
        })

        # Check for regressions/improvements at each stream count
        baseline_by_streams = {r.stream_count: r for r in baseline.results}
        test_by_streams = {r.stream_count: r for r in test.results}

        common_streams = set(baseline_by_streams.keys()) & set(test_by_streams.keys())

        for sc in sorted(common_streams):
            b_result = baseline_by_streams[sc]
            t_result = test_by_streams[sc]

            # Throughput comparison (higher is better)
            if b_result.throughput > 0:
                tp_change = (t_result.throughput - b_result.throughput) / b_result.throughput
                if tp_change < -threshold:
                    regressions.append({
                        "Workload": workload,
                        "Stream_Count": sc,
                        "Metric": "throughput",
                        "Baseline": round(b_result.throughput, 2),
                        "Test": round(t_result.throughput, 2),
                        "Change_%": round(tp_change * 100, 2),
                    })
                elif tp_change > threshold:
                    improvements.append({
                        "Workload": workload,
                        "Stream_Count": sc,
                        "Metric": "throughput",
                        "Baseline": round(b_result.throughput, 2),
                        "Test": round(t_result.throughput, 2),
                        "Change_%": round(tp_change * 100, 2),
                    })

            # P99 latency comparison (lower is better)
            if b_result.latency.p99 > 0:
                lat_change = (t_result.latency.p99 - b_result.latency.p99) / b_result.latency.p99
                if lat_change > threshold:  # Higher latency is regression
                    regressions.append({
                        "Workload": workload,
                        "Stream_Count": sc,
                        "Metric": "latency_p99",
                        "Baseline": round(b_result.latency.p99, 3),
                        "Test": round(t_result.latency.p99, 3),
                        "Change_%": round(lat_change * 100, 2),
                    })
                elif lat_change < -threshold:  # Lower latency is improvement
                    improvements.append({
                        "Workload": workload,
                        "Stream_Count": sc,
                        "Metric": "latency_p99",
                        "Baseline": round(b_result.latency.p99, 3),
                        "Test": round(t_result.latency.p99, 3),
                        "Change_%": round(lat_change * 100, 2),
                    })

    return summary_rows, regressions, improvements


def _build_throughput_by_streams_sheet(
    baseline_data: Dict[str, SweepData],
    test_data: Dict[str, SweepData],
    common_workloads: List[str],
) -> pd.DataFrame:
    """Build throughput comparison by stream count."""
    # Collect all stream counts
    all_streams = set()
    for wl in common_workloads:
        for r in baseline_data[wl].results:
            all_streams.add(r.stream_count)
        for r in test_data[wl].results:
            all_streams.add(r.stream_count)

    sorted_streams = sorted(all_streams)

    rows = []
    for wl in common_workloads:
        row = {"Workload": wl}

        b_by_streams = {r.stream_count: r.throughput for r in baseline_data[wl].results}
        t_by_streams = {r.stream_count: r.throughput for r in test_data[wl].results}

        for sc in sorted_streams:
            b_val = b_by_streams.get(sc)
            t_val = t_by_streams.get(sc)

            row[f"{sc}_Base"] = round(b_val, 2) if b_val else None
            row[f"{sc}_Test"] = round(t_val, 2) if t_val else None

            if b_val and t_val and b_val > 0:
                change = (t_val - b_val) / b_val * 100
                row[f"{sc}_Δ%"] = round(change, 1)
            else:
                row[f"{sc}_Δ%"] = None

        rows.append(row)

    return pd.DataFrame(rows)


def _build_latency_by_streams_sheet(
    baseline_data: Dict[str, SweepData],
    test_data: Dict[str, SweepData],
    common_workloads: List[str],
) -> pd.DataFrame:
    """Build P99 latency comparison by stream count."""
    # Collect all stream counts
    all_streams = set()
    for wl in common_workloads:
        for r in baseline_data[wl].results:
            all_streams.add(r.stream_count)
        for r in test_data[wl].results:
            all_streams.add(r.stream_count)

    sorted_streams = sorted(all_streams)

    rows = []
    for wl in common_workloads:
        row = {"Workload": wl}

        b_by_streams = {r.stream_count: r.latency.p99 for r in baseline_data[wl].results}
        t_by_streams = {r.stream_count: r.latency.p99 for r in test_data[wl].results}

        for sc in sorted_streams:
            b_val = b_by_streams.get(sc)
            t_val = t_by_streams.get(sc)

            row[f"{sc}_Base"] = round(b_val, 3) if b_val is not None else None
            row[f"{sc}_Test"] = round(t_val, 3) if t_val is not None else None

            if b_val is not None and t_val is not None and b_val != 0:
                change = (t_val - b_val) / b_val * 100
                row[f"{sc}_Δ%"] = round(change, 1)
            else:
                row[f"{sc}_Δ%"] = None

        rows.append(row)

    return pd.DataFrame(rows)


def _build_environment_comparison_sheet(
    baseline_data: Dict[str, SweepData],
    test_data: Dict[str, SweepData],
    baseline_label: str,
    test_label: str,
) -> pd.DataFrame:
    """Build environment comparison sheet."""
    # Get first workload's environment for comparison
    first_wl = next(iter(baseline_data.keys()))
    b_env = baseline_data[first_wl].environment
    t_env = test_data[first_wl].environment

    # Get GPU model from gpus list if available
    b_gpu_model = b_env.gpus[0] if b_env.gpus else "N/A"
    t_gpu_model = t_env.gpus[0] if t_env.gpus else "N/A"

    rows = [
        {"Property": "Label", baseline_label: baseline_label, test_label: test_label},
        {
            "Property": "Hostname",
            baseline_label: b_env.hostname if b_env.hostname is not None else "N/A",
            test_label: t_env.hostname if t_env.hostname is not None else "N/A",
        },
        {
            "Property": "GPU_Count",
            baseline_label: b_env.gpu_count if b_env.gpu_count is not None else "N/A",
            test_label: t_env.gpu_count if t_env.gpu_count is not None else "N/A",
        },
        {"Property": "GPU_Model", baseline_label: b_gpu_model, test_label: t_gpu_model},
        {
            "Property": "HIP_Version",
            baseline_label: b_env.hip_version if b_env.hip_version is not None else "N/A",
            test_label: t_env.hip_version if t_env.hip_version is not None else "N/A",
        },
        {
            "Property": "PyTorch_Version",
            baseline_label: b_env.torch_version if b_env.torch_version is not None else "N/A",
            test_label: t_env.torch_version if t_env.torch_version is not None else "N/A",
        },
        {
            "Property": "Driver_Type",
            baseline_label: b_env.driver_type if b_env.driver_type is not None else "N/A",
            test_label: t_env.driver_type if t_env.driver_type is not None else "N/A",
        },
        {
            "Property": "Kernel",
            baseline_label: b_env.kernel if b_env.kernel is not None else "N/A",
            test_label: t_env.kernel if t_env.kernel is not None else "N/A",
        },
    ]

    return pd.DataFrame(rows)


def generate_comparison_excel(
    baseline_data: Dict[str, SweepData],
    test_data: Dict[str, SweepData],
    common_workloads: List[str],
    baseline_only: List[str],
    test_only: List[str],
    output_path: Path,
    baseline_label: str = "Baseline",
    test_label: str = "Test",
    threshold: float = 0.05,
    verbose: bool = False,
) -> Tuple[Path, List[Dict], List[Dict]]:
    """
    Generate Excel report for multi-workload comparison (Mode C).

    Args:
        baseline_data: Dict of workload_name -> SweepData for baseline
        test_data: Dict of workload_name -> SweepData for test
        common_workloads: List of workloads in both baseline and test
        baseline_only: List of workloads only in baseline
        test_only: List of workloads only in test
        output_path: Path to output Excel file
        baseline_label: Label for baseline data
        test_label: Label for test data
        threshold: Regression threshold (fraction)
        verbose: Print verbose output

    Returns:
        Tuple of (output_path, regressions_list, improvements_list)
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if verbose:
        print(f"  Generating comparison Excel: {output_path.name}")
        print(f"    Workloads: {len(common_workloads)} common")
        print(f"    Threshold: {threshold * 100:.1f}%")

    # Compute comparison metrics
    summary_rows, regressions, improvements = _compute_comparison_metrics(
        baseline_data, test_data, common_workloads, threshold
    )

    # Build sheets
    sheets = {}

    # 1. Summary sheet
    summary_df = pd.DataFrame(summary_rows)
    sheets["Summary"] = summary_df

    # 2. Throughput by stream count
    throughput_df = _build_throughput_by_streams_sheet(
        baseline_data, test_data, common_workloads
    )
    sheets["Throughput_by_Streams"] = throughput_df

    # 3. Latency P99 by stream count
    latency_df = _build_latency_by_streams_sheet(
        baseline_data, test_data, common_workloads
    )
    sheets["Latency_P99_by_Streams"] = latency_df

    # 4. Regressions sheet
    if regressions:
        sheets["Regressions"] = pd.DataFrame(regressions)
    else:
        sheets["Regressions"] = pd.DataFrame(
            [{"Note": "No regressions detected"}]
        )

    # 5. Improvements sheet
    if improvements:
        sheets["Improvements"] = pd.DataFrame(improvements)
    else:
        sheets["Improvements"] = pd.DataFrame(
            [{"Note": "No significant improvements detected"}]
        )

    # 6. Missing workloads sheet
    missing_rows = []
    for wl in baseline_only:
        missing_rows.append({
            "Workload": wl,
            "Present_In": "Baseline",
            "Missing_From": "Test",
        })
    for wl in test_only:
        missing_rows.append({
            "Workload": wl,
            "Present_In": "Test",
            "Missing_From": "Baseline",
        })

    if missing_rows:
        sheets["Missing_Workloads"] = pd.DataFrame(missing_rows)
    else:
        sheets["Missing_Workloads"] = pd.DataFrame(
            [{"Note": "All workloads present in both baseline and test"}]
        )

    # 7. Environment comparison
    if common_workloads:
        env_df = _build_environment_comparison_sheet(
            baseline_data, test_data, baseline_label, test_label
        )
        sheets["Environment"] = env_df

    # Write to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply formatting with color-coded change columns
    change_columns = [
        ("Summary", "Change_%"),
        ("Throughput_by_Streams", "Δ%"),
        ("Latency_P99_by_Streams", "Δ%"),
        ("Regressions", "Change_%"),
        ("Improvements", "Change_%"),
    ]
    _apply_comparison_formatting(output_path, change_columns, verbose)

    if verbose:
        print(f"    Created {len(sheets)} sheet(s)")
        print(f"    Regressions found: {len(regressions)}")
        print(f"    Improvements found: {len(improvements)}")

    return output_path, regressions, improvements

