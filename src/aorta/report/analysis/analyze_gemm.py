"""
Analyze GEMM reports from TraceLens Excel files.

Extracts top N kernels with largest time variance (max - min) from
GEMM sheet data in individual performance reports.
"""

import csv
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import openpyxl


def extract_name_from_kernel_info(kernel_info_str: str) -> Optional[str]:
    """
    Extract the 'name' field from the kernel info string.

    Args:
        kernel_info_str: String containing kernel details, e.g.,
                        "[{'name': '...', 'stream': ..., ...}]"

    Returns:
        Kernel name or None if extraction fails
    """
    try:
        if kernel_info_str is None or kernel_info_str == "":
            return None

        # Try to extract just the name using regex
        match = re.search(r"'name':\s*'([^']+)'", str(kernel_info_str))
        if match:
            return match.group(1)

        return None
    except Exception:
        return None


def column_letter_to_index(letter: str) -> int:
    """Convert Excel column letter to 0-based index."""
    index = 0
    for i, char in enumerate(reversed(letter.upper())):
        index += (ord(char) - ord("A") + 1) * (26**i)
    return index - 1


def process_excel_file(
    file_path: Path,
    threads: int,
    channel: int,
    rank: int,
    top_k: int = 5,
) -> List[Dict[str, Any]]:
    """
    Process a single Excel file and extract GEMM data.

    Args:
        file_path: Path to the Excel file
        threads: Thread configuration
        channel: Channel configuration
        rank: Rank number
        top_k: Number of top kernels to extract

    Returns:
        List of dictionaries containing kernel data
    """
    try:
        # Open the workbook
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Check if GEMM sheet exists
        if "GEMM" not in wb.sheetnames:
            print(f"Warning: GEMM sheet not found in {file_path}")
            return []

        sheet = wb["GEMM"]

        # Expected column positions (0-based indices)
        col_kernel_info = column_letter_to_index("Y")  # Column X
        col_time_min = column_letter_to_index("AH")  # Column AG
        col_time_max = column_letter_to_index("AI")  # Column AH

        # Read header row to validate column names
        rows_data = []
        header_row = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                # This is the header - validate column names match expectations
                header_row = list(row)

                # Expected column names (match what TraceLens generates)
                expected_x = "kernel_details__summarize_kernel_stats"
                expected_ag = "Kernel Time (µs)_min"
                expected_ah = "Kernel Time (µs)_max"

                # Validate each expected column
                errors = []

                if col_kernel_info < len(header_row):
                    header_x = str(header_row[col_kernel_info]) if header_row[col_kernel_info] else ""
                    if header_x != expected_x:
                        errors.append(f"Column X: expected '{expected_x}', found '{header_x}'")
                else:
                    errors.append(f"Column X: not found (only {len(header_row)} columns)")

                if col_time_min < len(header_row):
                    header_ag = str(header_row[col_time_min]) if header_row[col_time_min] else ""
                    if header_ag != expected_ag:
                        errors.append(f"Column AG: expected '{expected_ag}', found '{header_ag}'")
                else:
                    errors.append(f"Column AG: not found (only {len(header_row)} columns)")

                if col_time_max < len(header_row):
                    header_ah = str(header_row[col_time_max]) if header_row[col_time_max] else ""
                    if header_ah != expected_ah:
                        errors.append(f"Column AH: expected '{expected_ah}', found '{header_ah}'")
                else:
                    errors.append(f"Column AH: not found (only {len(header_row)} columns)")

                if errors:
                    raise ValueError(
                        f"Column validation failed in {file_path}:\n  " + "\n  ".join(errors)
                    )

                continue

            if row is None or len(row) <= max(col_kernel_info, col_time_min, col_time_max):
                continue

            kernel_info = row[col_kernel_info] if col_kernel_info < len(row) else None
            kernel_time_min = row[col_time_min] if col_time_min < len(row) else None
            kernel_time_max = row[col_time_max] if col_time_max < len(row) else None

            # Extract kernel name
            kernel_name = extract_name_from_kernel_info(kernel_info)

            # Calculate time difference
            if kernel_time_min is not None and kernel_time_max is not None:
                try:
                    time_diff = float(kernel_time_max) - float(kernel_time_min)
                except (ValueError, TypeError):
                    continue
            else:
                continue

            if kernel_name:
                row_dict = {
                    "threads": threads,
                    "channel": channel,
                    "rank": rank,
                    "kernel_name": kernel_name,
                    "kernel_time_min_us": kernel_time_min,
                    "kernel_time_max_us": kernel_time_max,
                    "time_diff_us": time_diff,
                }

                # Add all other columns
                if header_row:
                    for j, val in enumerate(row):
                        if j < len(header_row) and header_row[j]:
                            col_name = f"col_{header_row[j]}"
                            row_dict[col_name] = val

                rows_data.append(row_dict)

        wb.close()

        # Sort by time_diff_us and get top k
        rows_data.sort(key=lambda x: x["time_diff_us"], reverse=True)
        top_results = rows_data[:top_k]

        return top_results

    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        import traceback
        traceback.print_exc()
        return []


def analyze_gemm_reports(
    base_path: Path,
    threads: List[int],
    channels: List[int],
    ranks: List[int],
    top_k: int = 5,
    output_file: str = "top5_gemm_kernels_time_variance.csv",
    verbose: bool = False,
) -> Optional[Path]:
    """
    Analyze GEMM reports from a sweep directory structure.

    Args:
        base_path: Path to tracelens_analysis directory
        threads: List of thread configurations to analyze (e.g., [256, 512])
        channels: List of channel configurations (e.g., [28, 42, 56, 70])
        ranks: List of ranks to analyze (e.g., [0, 1, 2, ..., 7])
        top_k: Number of top kernels to extract per file
        output_file: Output CSV filename
        verbose: Whether to print verbose output

    Returns:
        Path to output file or None if no data processed
    """
    # Validate base path
    if not base_path.exists():
        raise FileNotFoundError(f"Base path does not exist: {base_path}")

    if verbose:
        print(f"Configuration:")
        print(f"  Base path: {base_path}")
        print(f"  Threads: {threads}")
        print(f"  Channels: {channels}")
        print(f"  Ranks: {ranks}")
        print(f"  Top K: {top_k}")
        print(f"  Output file: {output_file}")
        print()

    all_results = []

    print("Processing Excel files...")
    total_files = len(threads) * len(channels) * len(ranks)
    file_count = 0

    for thread_count in threads:
        thread_dir = base_path / f"{thread_count}thread" / "individual_reports"

        for channel in channels:
            for rank in ranks:
                file_name = f"perf_{channel}ch_rank{rank}.xlsx"
                file_path = thread_dir / file_name

                file_count += 1
                if verbose:
                    print(f"Processing {file_count}/{total_files}: {file_name}")

                if not file_path.exists():
                    if verbose:
                        print(f"  Warning: File not found: {file_path}")
                    continue

                # Process the file
                results = process_excel_file(file_path, thread_count, channel, rank, top_k)

                if results:
                    all_results.extend(results)
                    if verbose:
                        print(f"  Found {len(results)} kernels")

    if not all_results:
        print("Error: No data extracted!")
        return None

    # Sort by time_diff_us descending
    print("\nCombining and sorting results...")
    all_results.sort(key=lambda x: x["time_diff_us"], reverse=True)

    # Get all unique keys
    all_keys = set()
    for row in all_results:
        all_keys.update(row.keys())

    # Order columns: metadata first, then others
    metadata_cols = [
        "threads",
        "channel",
        "rank",
        "kernel_name",
        "kernel_time_min_us",
        "kernel_time_max_us",
        "time_diff_us",
    ]
    other_cols = sorted([k for k in all_keys if k not in metadata_cols])
    ordered_cols = metadata_cols + other_cols

    # Determine output path
    output_path = Path(output_file)
    if not output_path.is_absolute():
        output_path = base_path / output_file

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save to CSV
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ordered_cols)
        writer.writeheader()

        for row in all_results:
            # Fill in missing keys with None
            full_row = {k: row.get(k, None) for k in ordered_cols}
            writer.writerow(full_row)

    print(f"\nResults saved to: {output_path}")
    print(f"Total rows: {len(all_results)}")

    # Print summary
    print(f"\nTop {min(10, len(all_results))} kernels by time difference:")
    for i, row in enumerate(all_results[:10]):
        print(
            f"{i+1}. threads={row['threads']}, ch={row['channel']}, rank={row['rank']}, "
            f"diff={row['time_diff_us']:.4f}us"
        )
        print(f"   {row['kernel_name'][:100]}...")

    # Print summary statistics
    time_diffs = [r["time_diff_us"] for r in all_results]
    kernel_names = set(r["kernel_name"] for r in all_results)

    print(f"\nSummary Statistics:")
    print(f"Total unique kernels: {len(kernel_names)}")
    print(f"Average time difference: {sum(time_diffs)/len(time_diffs):.4f} us")
    print(f"Max time difference: {max(time_diffs):.4f} us")
    print(f"Min time difference: {min(time_diffs):.4f} us")

    return output_path

