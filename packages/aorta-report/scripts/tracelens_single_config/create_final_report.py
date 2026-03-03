#!/usr/bin/env python3
"""
Create final comprehensive report with combined and comparison data.
Raw data sheets are hidden and all data is formatted as Excel tables.
"""
import pandas as pd
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color
from openpyxl.formatting.rule import ColorScaleRule


def get_column_letter(col_num):
    """Convert column number to Excel column letter."""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result


def add_excel_table(worksheet, table_name, start_row=1):
    """Convert worksheet data to Excel table format."""
    # Find data range
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    if max_row <= start_row:
        return  # No data

    # Ensure all column headers are strings
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=start_row, column=col_idx)
        if cell.value is not None and not isinstance(cell.value, str):
            cell.value = str(cell.value)

    # Create table reference using proper column letter conversion
    start_cell = f"A{start_row}"
    end_col_letter = get_column_letter(max_col)
    end_cell = f"{end_col_letter}{max_row}"
    table_ref = f"{start_cell}:{end_cell}"

    # Create table with style
    try:
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style

        # Add table to worksheet
        worksheet.add_table(tab)
    except Exception as e:
        print(f"    Warning: Could not create table {table_name}: {e}")


def create_final_report(
    gpu_combined, gpu_comparison, coll_combined, coll_comparison, output_file,baseline_label='Baseline', test_label='Test'
):
    """Create comprehensive report with all data."""

    print("Creating comprehensive final report...")
    print(f"  Output: {output_file}")

    # Track sheet info for hiding/organizing
    raw_sheets = []
    comparison_sheets = []
    summary_sheets = []

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # === GPU TIMELINE SHEETS ===
        print("\nAdding GPU Timeline sheets...")

        # Read GPU combined (raw data)
        gpu_comb_xl = pd.ExcelFile(gpu_combined)
        sheet_mapping = {
            "Summary": "GPU_Summary_Raw",
            "All_Ranks_Combined": "GPU_AllRanks_Raw",
            "Per_Rank_Time_ms": "GPU_Time_Raw",
            "Per_Rank_Percent": "GPU_Pct_Raw",
        }
        for sheet_name in gpu_comb_xl.sheet_names:
            df = pd.read_excel(gpu_combined, sheet_name=sheet_name)
            new_name = sheet_mapping.get(sheet_name, f"GPU_{sheet_name}_Raw")
            df.to_excel(writer, sheet_name=new_name, index=False)
            raw_sheets.append(new_name)
            print(f"  Added {new_name} (will be hidden)")

        # Read GPU comparison
        gpu_comp_xl = pd.ExcelFile(gpu_comparison)
        comp_mapping = {
            "Summary_Comparison": "GPU_Summary_Cmp",
            "Comparison_By_Rank": "GPU_ByRank_Cmp",
        }
        for sheet_name in gpu_comp_xl.sheet_names:
            if "Comparison" in sheet_name:
                df = pd.read_excel(gpu_comparison, sheet_name=sheet_name)
                new_name = comp_mapping.get(sheet_name, f"GPU_{sheet_name}")
                df.to_excel(writer, sheet_name=new_name, index=False)
                comparison_sheets.append(new_name)
                print(f"  Added {new_name}")

        # === COLLECTIVE SHEETS ===
        print("\nAdding Collective/NCCL sheets...")

        # Read collective combined (raw data for hidden sheets)
        coll_comb_xl = pd.ExcelFile(coll_combined)
        coll_mapping = {
            "nccl_summary_implicit_sync": "NCCL_ImplSync_Raw",
            "nccl_summary_long": "NCCL_Long_Raw",
        }
        for sheet_name in coll_comb_xl.sheet_names:
            if "summary" in sheet_name.lower():
                df = pd.read_excel(coll_combined, sheet_name=sheet_name)
                new_name = coll_mapping.get(sheet_name, f"NCCL_{sheet_name}_Raw")
                df.to_excel(writer, sheet_name=new_name, index=False)
                raw_sheets.append(new_name)
                print(f"  Added {new_name} (will be hidden)")

        # Read collective comparison
        coll_comp_xl = pd.ExcelFile(coll_comparison)
        for sheet_name in coll_comp_xl.sheet_names:
            df = pd.read_excel(coll_comparison, sheet_name=sheet_name)

            # Determine appropriate naming
            if 'nccl' in sheet_name.lower():
                if '_cmp' in sheet_name or 'comparison' in sheet_name.lower():
                    new_name = f"NCCL_{sheet_name.replace('nccl_', '').title().replace('_', '')}"
                else:
                    new_name = f"NCCL_{sheet_name}"
            else:
                new_name = sheet_name

            df.to_excel(writer, sheet_name=new_name, index=False)

            if '_cmp' in sheet_name.lower() or 'comparison' in sheet_name.lower():
                comparison_sheets.append(new_name)
            else:
                raw_sheets.append(new_name)

            print(f"  Added {new_name}")

        # === CREATE SUMMARY DASHBOARD ===
        print("\nCreating Summary Dashboard...")

        # Read key metrics for dashboard
        gpu_summary = pd.read_excel(gpu_comparison, sheet_name="Summary_Comparison")

        # Create dashboard data
        dashboard_data = {
            'Metric': [],
            baseline_label: [],
            test_label: [],
            'Improvement (%)': [],
            'Status': []
        }

        # Add GPU metrics
        # Find the actual column names (they may be config-specific like '32cu_512threads_time_ms')
        time_cols = [col for col in gpu_summary.columns if 'time_ms' in col and 'diff' not in col and 'percent' not in col]
        if len(time_cols) >= 2:
            baseline_col = time_cols[0]
            test_col = time_cols[1]
        else:
            # Fallback to default names
            baseline_col = 'baseline_time_ms' if 'baseline_time_ms' in gpu_summary.columns else time_cols[0] if time_cols else None
            test_col = 'test_time_ms' if 'test_time_ms' in gpu_summary.columns else time_cols[1] if len(time_cols) > 1 else None

        if baseline_col and test_col:
            for _, row in gpu_summary.iterrows():
                metric_type = row['type']
                dashboard_data['Metric'].append(f"GPU_{metric_type}")
                dashboard_data[baseline_label].append(round(row[baseline_col], 2))
                dashboard_data[test_label].append(round(row[test_col], 2))
                dashboard_data['Improvement (%)'].append(round(row['percent_change'], 2) if 'percent_change' in row else 0)

                pct_val = row['percent_change'] if 'percent_change' in row else 0
                dashboard_data['Status'].append('Better' if pct_val > 0 else 'Worse' if pct_val < -1 else 'Similar')

        dashboard_df = pd.DataFrame(dashboard_data)
        dashboard_df.to_excel(writer, sheet_name="Summary_Dashboard", index=False)
        summary_sheets.append("Summary_Dashboard")
        print(f"  Added Summary_Dashboard")

    # Now modify the workbook to hide sheets and add tables
    print("\nApplying formatting...")
    wb = load_workbook(output_file)

    # Hide raw data sheets
    for sheet_name in raw_sheets:
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_state = "hidden"
            print(f"  Hidden: {sheet_name}")

    # Convert all sheets to tables
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip if sheet is empty
        if ws.max_row <= 1:
            continue

        # Create unique table name from sheet name (remove special chars)
        table_name = (
            sheet_name.replace(" ", "_")
            .replace("-", "_")
            .replace("(", "")
            .replace(")", "")
        )
        # Ensure name starts with letter and is max 255 chars
        if not table_name[0].isalpha():
            table_name = "Tbl_" + table_name
        table_name = table_name[:255]

        add_excel_table(ws, table_name)
        print(f"  Converted to table: {sheet_name}")

        # Add conditional formatting for percent_change columns
        if "Cmp" in sheet_name or "Comparison" in sheet_name:
            # Find percent_change columns
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value and "percent_change" in str(cell_value):
                    col_letter = get_column_letter(col_idx)
                    data_range = f"{col_letter}2:{col_letter}{ws.max_row}"

                    # Apply color scale: red (min/negative) -> white (0) -> green (max/positive)
                    try:
                        ws.conditional_formatting.add(
                            data_range,
                            ColorScaleRule(
                                start_type="min",
                                start_color="F8696B",  # Red
                                mid_type="num",
                                mid_value=0,
                                mid_color="FFFFFF",  # White
                                end_type="max",
                                end_color="63BE7B",  # Green
                            ),
                        )
                        print(
                            f"    Applied color scale to {sheet_name} column {cell_value}"
                        )
                    except Exception as e:
                        print(
                            f"    Warning: Could not apply formatting to {cell_value}: {e}"
                        )

    # Move Summary Dashboard to first position
    if "Summary_Dashboard" in wb.sheetnames:
        dashboard_sheet = wb["Summary_Dashboard"]
        wb.move_sheet(dashboard_sheet, offset=-(len(wb.sheetnames) - 1))
        wb.active = 0  # Set dashboard as active sheet
        print("\n  Moved Summary_Dashboard to first position")

    # Save workbook
    wb.save(output_file)
    print(f"\nFinal report saved: {output_file}")

    # Report structure
    print("\nReport Structure:")
    print("  Visible Sheets (Analysis):")
    print(f"    - Summary_Dashboard")
    for sheet in comparison_sheets:
        print(f"    - {sheet}")
    print("\n  Hidden Sheets (Raw Data):")
    for sheet in raw_sheets:
        print(f"    - {sheet}")
    print("\n  All data formatted as Excel tables with filters")
    print("  Percent change columns are color-coded (green=better, red=worse)")
    print(
        "\nUsers can unhide raw data sheets in Excel: Right-click any sheet tab → Unhide"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Create final comprehensive report with all data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example:
  python create_final_report.py \\
    --gpu-combined gpu_timeline_combined.xlsx \\
    --gpu-comparison gpu_timeline_comparison.xlsx \\
    --coll-combined collective_combined.xlsx \\
    --coll-comparison collective_comparison.xlsx \\
    --output final_analysis_report.xlsx
        """,
    )

    parser.add_argument(
        "--gpu-combined", required=True, help="Path to GPU timeline combined file"
    )
    parser.add_argument(
        "--gpu-comparison", required=True, help="Path to GPU timeline comparison file"
    )
    parser.add_argument(
        "--coll-combined", required=True, help="Path to collective combined file"
    )
    parser.add_argument(
        "--coll-comparison", required=True, help="Path to collective comparison file"
    )
    parser.add_argument("--output", required=True, help="Output path for final report")
    parser.add_argument('--baseline-label', default='Baseline',
                       help='Label for baseline configuration')
    parser.add_argument('--test-label', default='Test',
                       help='Label for test configuration')

    args = parser.parse_args()

    # Validate inputs
    for file_arg in [
        "gpu_combined",
        "gpu_comparison",
        "coll_combined",
        "coll_comparison",
    ]:
        file_path = getattr(args, file_arg)
        if not Path(file_path).exists():
            print(f"Error: File not found: {file_path}")
            return 1

    create_final_report(
        args.gpu_combined,
        args.gpu_comparison,
        args.coll_combined,
        args.coll_comparison,
        args.output,
        args.baseline_label,
        args.test_label
    )

    return 0


if __name__ == "__main__":
    exit(main())
