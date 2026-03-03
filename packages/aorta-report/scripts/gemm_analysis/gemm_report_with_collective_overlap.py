#!/usr/bin/env python3
"""
Enhance the GEMM variance CSV with collective overlap information.
For each row, check if any NCCL collective operations overlapped with the max duration GEMM kernel.
"""

import os
os.environ['OPENBLAS_NUM_THREADS'] = '1'
os.environ['MKL_NUM_THREADS'] = '1'

import pandas as pd
import openpyxl
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple

def calculate_unique_overlap_duration(intervals: List[Tuple[float, float]]) -> float:
    """
    Calculate total unique duration from potentially overlapping intervals.
    Uses interval merging to avoid double-counting overlapping time periods.

    Args:
        intervals: List of (start, end) tuples representing time intervals

    Returns:
        Total unique duration covered by all intervals
    """
    if not intervals:
        return 0.0

    # Sort intervals by start time
    sorted_intervals = sorted(intervals)

    # Merge overlapping intervals
    merged = []
    current_start, current_end = sorted_intervals[0]

    for start, end in sorted_intervals[1:]:
        if start <= current_end:
            # Overlapping or adjacent - merge
            current_end = max(current_end, end)
        else:
            # Non-overlapping - save current and start new
            merged.append((current_start, current_end))
            current_start, current_end = start, end

    # Don't forget the last interval
    merged.append((current_start, current_end))

    # Calculate total duration
    total_duration = sum(end - start for start, end in merged)

    return total_duration

def load_collective_data(collective_file: Path) -> pd.DataFrame:
    """Load NCCL collective data from Excel file."""
    try:
        wb = openpyxl.load_workbook(collective_file, read_only=True, data_only=True)

        if 'nccl_long' not in wb.sheetnames:
            print(f"Warning: 'nccl_long' sheet not found in {collective_file}")
            return pd.DataFrame()

        sheet = wb['nccl_long']

        # Extract data from relevant columns
        data = []
        headers = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = row
                continue

            if row and len(row) > 16:  # Ensure we have enough columns
                # Column indices (0-based):
                # C = 2 (rank)
                # F = 5 (collective name)
                # P = 15 (timestamp)
                # Q = 16 (duration)

                data.append({
                    'rank': row[2],
                    'collective_name': row[5],
                    'timestamp_ms': row[15] / 1000.0 if row[15] else None,  # Convert us to ms
                    'duration_ms': row[16] / 1000.0 if row[16] else None  # Convert us to ms
                })

        wb.close()

        df = pd.DataFrame(data)

        # Clean data
        if not df.empty:
            # Remove rows with missing critical values
            df = df.dropna(subset=['rank', 'timestamp_ms', 'duration_ms'])

            # Calculate end timestamp
            df['end_timestamp_ms'] = df['timestamp_ms'] + df['duration_ms']

            # Convert rank to int
            df['rank'] = df['rank'].astype(int)

        return df

    except Exception as e:
        print(f"Error loading collective data from {collective_file}: {e}")
        return pd.DataFrame()

def find_overlapping_collectives(gemm_timestamp_ms: float, gemm_duration_us: float,
                               collectives_df: pd.DataFrame, rank: int) -> Tuple[List[Dict], float]:
    """
    Find collective operations that overlap with a GEMM kernel.

    Args:
        gemm_timestamp_ms: Start timestamp of GEMM kernel in milliseconds
        gemm_duration_us: Duration of GEMM kernel in microseconds
        collectives_df: DataFrame with collective operations
        rank: Rank to filter collectives for

    Returns:
        Tuple of (list of overlapping collective operations, total unique overlap duration in ms)
    """
    if collectives_df.empty or pd.isna(gemm_timestamp_ms):
        return [], 0.0

    # Convert GEMM duration to milliseconds and calculate end time
    gemm_duration_ms = gemm_duration_us / 1000.0
    gemm_end_ms = gemm_timestamp_ms + gemm_duration_ms

    # Filter collectives for this rank
    rank_collectives = collectives_df[collectives_df['rank'] == rank]

    # Find overlapping collectives
    # Overlap occurs when: collective_start < gemm_end AND collective_end > gemm_start
    overlapping = rank_collectives[
        (rank_collectives['timestamp_ms'] < gemm_end_ms) &
        (rank_collectives['end_timestamp_ms'] > gemm_timestamp_ms)
    ]

    # Convert to list of dicts with overlap info
    overlaps = []
    overlap_intervals = []  # Store (start, end) tuples for unique time calculation

    for _, coll in overlapping.iterrows():
        # Calculate overlap duration
        overlap_start = max(gemm_timestamp_ms, coll['timestamp_ms'])
        overlap_end = min(gemm_end_ms, coll['end_timestamp_ms'])
        overlap_duration_ms = overlap_end - overlap_start

        overlaps.append({
            'collective_name': coll['collective_name'],
            'collective_timestamp_ms': coll['timestamp_ms'],
            'collective_duration_ms': coll['duration_ms'],
            'overlap_duration_ms': overlap_duration_ms,
            'overlap_percentage': (overlap_duration_ms / gemm_duration_ms) * 100
        })

        overlap_intervals.append((overlap_start, overlap_end))

    # Calculate total unique overlap duration (merge overlapping intervals)
    unique_overlap_ms = calculate_unique_overlap_duration(overlap_intervals)

    return overlaps, unique_overlap_ms

def enhance_csv_with_collective_overlap(input_csv: Path, output_csv: Path,
                                      tracelens_path: Path):
    """Add collective overlap information to the GEMM variance CSV."""

    # Read the input CSV
    df = pd.read_csv(input_csv)

    # Add new columns for collective overlap info
    df['overlapping_collective_count'] = 0
    df['overlapping_collective_names'] = ''
    df['max_overlap_percentage'] = 0.0
    df['max_overlap_collective'] = ''
    df['total_overlap_duration_ms'] = 0.0

    total_rows = len(df)
    print(f"Processing {total_rows} rows...")

    # Cache for collective data
    collective_cache = {}

    for idx, row in df.iterrows():
        print(f"\nProcessing row {idx + 1}/{total_rows}")

        threads = int(row['threads'])
        channel = int(row['channel'])
        rank = int(row['rank'])

        # Get max duration GEMM timestamp and duration
        max_timestamp_ms = row.get('max_duration_timestamp_ms')
        max_duration_us = row.get('max_duration_found_us', row.get('kernel_time_max_us'))

        if pd.isna(max_timestamp_ms) or pd.isna(max_duration_us):
            print(f"  Skipping - no timestamp or duration for max GEMM")
            continue

        print(f"  Config: {threads}thread/{channel}ch/rank{rank}")
        print(f"  Max GEMM: timestamp={max_timestamp_ms:.3f}ms, duration={max_duration_us:.3f}us")

        # Load collective data (cache it)
        cache_key = f"{threads}thread/{channel}ch"
        if cache_key not in collective_cache:
            collective_file = tracelens_path / f"{threads}thread/collective_reports/collective_{channel}ch.xlsx"

            if not collective_file.exists():
                print(f"  Warning: Collective file not found: {collective_file}")
                collective_cache[cache_key] = pd.DataFrame()
            else:
                print(f"  Loading collective data from: {collective_file.name}")
                collective_cache[cache_key] = load_collective_data(collective_file)

        collectives_df = collective_cache[cache_key]

        if collectives_df.empty:
            continue

        # Find overlapping collectives
        overlaps, unique_overlap_ms = find_overlapping_collectives(
            max_timestamp_ms, max_duration_us, collectives_df, rank
        )

        if overlaps:
            print(f"  Found {len(overlaps)} overlapping collective(s):")

            # Update DataFrame
            df.at[idx, 'overlapping_collective_count'] = len(overlaps)

            # Aggregate collective names
            collective_names = [o['collective_name'] for o in overlaps]
            df.at[idx, 'overlapping_collective_names'] = ';'.join(collective_names)

            # Find max overlap
            max_overlap = max(overlaps, key=lambda x: x['overlap_percentage'])
            df.at[idx, 'max_overlap_percentage'] = max_overlap['overlap_percentage']
            df.at[idx, 'max_overlap_collective'] = max_overlap['collective_name']

            # Total unique overlap duration (no double-counting)
            df.at[idx, 'total_overlap_duration_ms'] = unique_overlap_ms

            # Print details
            for overlap in overlaps:
                print(f"    - {overlap['collective_name']}: "
                      f"{overlap['overlap_percentage']:.1f}% overlap "
                      f"({overlap['overlap_duration_ms']:.3f}ms)")

            # Show summary
            summed_overlap = sum(o['overlap_duration_ms'] for o in overlaps)
            gemm_duration_ms = max_duration_us / 1000.0
            print(f"  Total: {unique_overlap_ms:.3f}ms unique overlap "
                  f"({unique_overlap_ms/gemm_duration_ms*100:.1f}% of GEMM duration)")
            if summed_overlap != unique_overlap_ms:
                print(f"  Note: Summed overlaps = {summed_overlap:.3f}ms "
                      f"(difference due to concurrent collectives)")
        else:
            print(f"  No overlapping collectives found")

    # Save enhanced CSV
    df.to_csv(output_csv, index=False)
    print(f"\nEnhanced CSV saved to: {output_csv}")

    # Print summary statistics
    rows_with_overlap = (df['overlapping_collective_count'] > 0).sum()
    print(f"\nSummary:")
    print(f"  Total rows: {total_rows}")
    print(f"  Rows with collective overlap: {rows_with_overlap}")
    print(f"  Overlap rate: {rows_with_overlap/total_rows*100:.1f}%")

    if rows_with_overlap > 0:
        # Analyze overlap patterns
        print(f"\nOverlap patterns:")
        collective_counts = df[df['overlapping_collective_count'] > 0]['overlapping_collective_names'].str.split(';', expand=True).stack().value_counts()
        print("  Most common overlapping collectives:")
        for coll, count in collective_counts.head(5).items():
            print(f"    - {coll}: {count} occurrences")

        # Overlap severity
        high_overlap = (df['max_overlap_percentage'] > 50).sum()
        print(f"\n  High overlap (>50%): {high_overlap} rows")

        avg_overlap = df[df['max_overlap_percentage'] > 0]['max_overlap_percentage'].mean()
        print(f"  Average max overlap: {avg_overlap:.1f}%")

def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Enhance GEMM variance CSV with collective overlap information",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument(
        '--input-csv',
        type=Path,
        required=True,
        help='Input CSV file with GEMM variance and timestamp data'
    )

    parser.add_argument(
        '--output-csv',
        type=Path,
        default=None,
        help='Output CSV file (default: input_file_with_collective_overlap.csv)'
    )

    parser.add_argument(
        '--tracelens-path',
        type=Path,
        required=True,
        help='Path to tracelens_analysis directory containing collective reports'
    )

    return parser.parse_args()

def main():
    args = parse_args()

    # Set default output file if not specified
    if args.output_csv is None:
        stem = args.input_csv.stem.replace('_with_timestamps', '')
        args.output_csv = args.input_csv.parent / f"{stem}_with_collective_overlap.csv"

    print("GEMM Variance Collective Overlap Analysis")
    print("=" * 60)
    print(f"Input CSV: {args.input_csv}")
    print(f"Output CSV: {args.output_csv}")
    print(f"TraceLens path: {args.tracelens_path}")
    print()

    # Verify input file exists
    if not args.input_csv.exists():
        print(f"Error: Input CSV not found: {args.input_csv}")
        return

    # Enhance the CSV with collective overlap info
    enhance_csv_with_collective_overlap(args.input_csv, args.output_csv, args.tracelens_path)

    print("\n[DONE] Enhancement complete!")

if __name__ == "__main__":
    main()
