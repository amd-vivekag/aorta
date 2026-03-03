#!/usr/bin/env python3
"""
Analyze GEMM reports from Excel files and extract top 5 kernels
with largest difference between max and min times.
"""

import re
import sys
import argparse
from pathlib import Path
import openpyxl
import csv

def extract_name_from_kernel_info(kernel_info_str):
    """Extract the 'name' field from the kernel info string."""
    try:
        # Parse the string to extract the name
        # The format is: "[{'name': '...', 'stream': ..., ...}]"
        if kernel_info_str is None or kernel_info_str == '':
            return None

        # Try to extract just the name using regex
        match = re.search(r"'name':\s*'([^']+)'", str(kernel_info_str))
        if match:
            return match.group(1)

        return None
    except Exception as e:
        print(f"Error parsing kernel info: {e}")
        return None

def column_letter_to_index(letter):
    """Convert Excel column letter to 0-based index."""
    index = 0
    for i, char in enumerate(reversed(letter.upper())):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index - 1

def process_excel_file(file_path, threads, channel, rank, top_k=5):
    """Process a single Excel file and extract GEMM data."""
    try:
        # Open the workbook
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Check if GEMM sheet exists
        if 'GEMM' not in wb.sheetnames:
            print(f"Warning: GEMM sheet not found in {file_path}")
            return []

        sheet = wb['GEMM']

        # Expected column positions (0-based indices)
        col_kernel_info = column_letter_to_index('X')    # Column X
        col_time_min = column_letter_to_index('AG')      # Column AG
        col_time_max = column_letter_to_index('AH')      # Column AH

        # Read header row to validate column names
        rows_data = []
        header_row = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                # This is the header - validate column names match expectations
                header_row = list(row)

                # Expected column names (match what TraceLens generates)
                expected_x = "kernel_details__summarize_kernel_stats"
                expected_ag = "Kernel Time (µs)_min"
                expected_ah = "Kernel Time (µs)_max"

                # Validate each expected column
                errors = []

                if col_kernel_info < len(header_row):
                    header_x = str(header_row[col_kernel_info]) if header_row[col_kernel_info] else ""
                    if header_x != expected_x:
                        errors.append(f"Column X: expected '{expected_x}', found '{header_x}'")
                else:
                    errors.append(f"Column X: not found (only {len(header_row)} columns)")

                if col_time_min < len(header_row):
                    header_ag = str(header_row[col_time_min]) if header_row[col_time_min] else ""
                    if header_ag != expected_ag:
                        errors.append(f"Column AG: expected '{expected_ag}', found '{header_ag}'")
                else:
                    errors.append(f"Column AG: not found (only {len(header_row)} columns)")

                if col_time_max < len(header_row):
                    header_ah = str(header_row[col_time_max]) if header_row[col_time_max] else ""
                    if header_ah != expected_ah:
                        errors.append(f"Column AH: expected '{expected_ah}', found '{header_ah}'")
                else:
                    errors.append(f"Column AH: not found (only {len(header_row)} columns)")

                if errors:
                    raise ValueError(
                        f"Column validation failed in {file_path}:\n  " +
                        "\n  ".join(errors)
                    )

                continue

            if row is None or len(row) <= max(col_kernel_info, col_time_min, col_time_max):
                continue

            kernel_info = row[col_kernel_info] if col_kernel_info < len(row) else None
            kernel_time_min = row[col_time_min] if col_time_min < len(row) else None
            kernel_time_max = row[col_time_max] if col_time_max < len(row) else None

            # Extract kernel name
            kernel_name = extract_name_from_kernel_info(kernel_info)

            # Calculate time difference
            if kernel_time_min is not None and kernel_time_max is not None:
                try:
                    time_diff = float(kernel_time_max) - float(kernel_time_min)
                except (ValueError, TypeError):
                    continue
            else:
                continue

            if kernel_name:
                row_dict = {
                    'threads': threads,
                    'channel': channel,
                    'rank': rank,
                    'kernel_name': kernel_name,
                    'kernel_time_min_us': kernel_time_min,
                    'kernel_time_max_us': kernel_time_max,
                    'time_diff_us': time_diff
                }

                # Add all other columns
                if header_row:
                    for j, val in enumerate(row):
                        if j < len(header_row) and header_row[j]:
                            col_name = f"col_{header_row[j]}"
                            row_dict[col_name] = val

                rows_data.append(row_dict)

        wb.close()

        # Sort by time_diff_us and get top k
        rows_data.sort(key=lambda x: x['time_diff_us'], reverse=True)
        top_results = rows_data[:top_k]

        return top_results

    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        import traceback
        traceback.print_exc()
        return []

def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Analyze GEMM reports and extract top kernels with highest time variance",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default settings
  python analyze_gemm_reports.py

  # Specify custom base path
  python analyze_gemm_reports.py --base-path /path/to/tracelens_analysis

  # Specify custom configurations
  python analyze_gemm_reports.py --threads 256 512 --channels 28 42 56 70 --ranks 0 1 2 3 4 5 6 7

  # Extract top 10 kernels instead of top 5
  python analyze_gemm_reports.py --top-k 10

  # Custom output file
  python analyze_gemm_reports.py --output-file my_analysis.csv
        """
    )

    parser.add_argument(
        '--base-path',
        type=Path,
        required=True,
        help='Base path to tracelens_analysis directory'
    )

    parser.add_argument(
        '--threads',
        type=int,
        nargs='+',
        default=[256, 512],
        help='Thread configurations to analyze (default: %(default)s)'
    )

    parser.add_argument(
        '--channels',
        type=int,
        nargs='+',
        default=[28, 42, 56, 70],
        help='Channel configurations to analyze (default: %(default)s)'
    )

    parser.add_argument(
        '--ranks',
        type=int,
        nargs='+',
        default=list(range(8)),
        help='Ranks to analyze (default: 0 1 2 3 4 5 6 7)'
    )

    parser.add_argument(
        '--top-k',
        type=int,
        default=5,
        help='Number of top kernels to extract per file (default: %(default)s)'
    )

    parser.add_argument(
        '--output-file',
        type=str,
        default='top5_gemm_kernels_time_variance.csv',
        help='Output CSV filename (default: %(default)s)'
    )

    return parser.parse_args()

def main():
    # Parse command line arguments
    args = parse_args()

    base_path = args.base_path
    thread_configs = args.threads
    channels = args.channels
    ranks = args.ranks
    top_k = args.top_k

    # Validate base path
    if not base_path.exists():
        print(f"Error: Base path does not exist: {base_path}")
        return

    print(f"Configuration:")
    print(f"  Base path: {base_path}")
    print(f"  Threads: {thread_configs}")
    print(f"  Channels: {channels}")
    print(f"  Ranks: {ranks}")
    print(f"  Top K: {top_k}")
    print(f"  Output file: {args.output_file}")
    print()

    all_results = []

    print("Processing Excel files...")
    total_files = len(thread_configs) * len(channels) * len(ranks)
    file_count = 0

    for threads in thread_configs:
        thread_dir = base_path / f"{threads}thread" / "individual_reports"

        for channel in channels:
            for rank in ranks:
                file_name = f"perf_{channel}ch_rank{rank}.xlsx"
                file_path = thread_dir / file_name

                file_count += 1
                print(f"Processing {file_count}/{total_files}: {file_name}")

                if not file_path.exists():
                    print(f"  Warning: File not found: {file_path}")
                    continue

                # Process the file
                results = process_excel_file(file_path, threads, channel, rank, top_k)

                if results:
                    all_results.extend(results)
                    print(f"  Found {len(results)} kernels")

    if not all_results:
        print("Error: No data extracted!")
        sys.exit(1)

    # Sort by time_diff_us descending
    print("\nCombining and sorting results...")
    all_results.sort(key=lambda x: x['time_diff_us'], reverse=True)

    # Get all unique keys
    all_keys = set()
    for row in all_results:
        all_keys.update(row.keys())

    # Order columns: metadata first, then others
    metadata_cols = ['threads', 'channel', 'rank', 'kernel_name', 'kernel_time_min_us', 'kernel_time_max_us', 'time_diff_us']
    other_cols = sorted([k for k in all_keys if k not in metadata_cols])
    ordered_cols = metadata_cols + other_cols

    # Save to CSV
    # Handle both absolute and relative paths
    output_file = Path(args.output_file)
    if not output_file.is_absolute():
        output_file = base_path / output_file

    # Ensure output directory exists
    try:
        output_file.parent.mkdir(parents=True, exist_ok=True)
    except (OSError, PermissionError) as e:
        print(f"Error: Cannot create output directory: {e}")
        sys.exit(1)

    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=ordered_cols)
            writer.writeheader()

            for row in all_results:
                # Fill in missing keys with None
                full_row = {k: row.get(k, None) for k in ordered_cols}
                writer.writerow(full_row)
    except (OSError, PermissionError, FileNotFoundError) as e:
        print(f"Error: Cannot write to output file {output_file}: {e}")
        sys.exit(1)

    print(f"\nResults saved to: {output_file}")
    print(f"Total rows: {len(all_results)}")

    print(f"\nTop {min(10, len(all_results))} kernels by time difference:")
    for i, row in enumerate(all_results[:10]):
        print(f"{i+1}. threads={row['threads']}, ch={row['channel']}, rank={row['rank']}, "
              f"diff={row['time_diff_us']:.4f}us")
        print(f"   {row['kernel_name'][:100]}...")

    # Print summary statistics
    time_diffs = [r['time_diff_us'] for r in all_results]
    kernel_names = set(r['kernel_name'] for r in all_results)

    print(f"\nSummary Statistics:")
    print(f"Total unique kernels: {len(kernel_names)}")
    print(f"Average time difference: {sum(time_diffs)/len(time_diffs):.4f} us")
    print(f"Max time difference: {max(time_diffs):.4f} us")
    print(f"Min time difference: {min(time_diffs):.4f} us")

if __name__ == "__main__":
    main()
