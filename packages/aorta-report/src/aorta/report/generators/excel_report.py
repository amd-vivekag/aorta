"""Final Excel report generator.

Creates comprehensive report with:
- Summary Dashboard (first, visible)
- Comparison sheets (visible)
- Raw data sheets (hidden)
- Excel table formatting
- Color-coded percent_change columns
"""

from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# Sheet Naming Mappings
# =============================================================================

GPU_SHEET_MAPPING = {
    "Summary": "GPU_Summary_Raw",
    "All_Ranks_Combined": "GPU_AllRanks_Raw",
    "Per_Rank_Time_ms": "GPU_Time_Raw",
    "Per_Rank_Percent": "GPU_Pct_Raw",
}

GPU_COMPARISON_MAPPING = {
    "Summary_Comparison": "GPU_Summary_Cmp",
    "Comparison_By_Rank": "GPU_ByRank_Cmp",
}

COLL_SHEET_MAPPING = {
    "nccl_summary_implicit_sync": "NCCL_ImplSync_Raw",
    "nccl_summary_long": "NCCL_Long_Raw",
}

# Color scale colors
RED = "F8696B"
WHITE = "FFFFFF"
GREEN = "63BE7B"


def sanitize_table_name(sheet_name: str) -> str:
    """Create valid Excel table name from sheet name."""
    table_name = sheet_name.replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
    # Ensure name starts with letter
    if not table_name[0].isalpha():
        table_name = "Tbl_" + table_name
    # Max 255 chars
    return table_name[:255]


def add_excel_table(worksheet: Worksheet, table_name: str, start_row: int = 1) -> bool:
    """Convert worksheet data to Excel table format.

    Returns True if table was added, False otherwise.
    """
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    if max_row <= start_row:
        return False  # No data

    # Ensure all column headers are strings
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=start_row, column=col_idx)
        if cell.value is not None and not isinstance(cell.value, str):
            cell.value = str(cell.value)

    # Create table reference
    start_cell = f"A{start_row}"
    end_col_letter = get_column_letter(max_col)
    end_cell = f"{end_col_letter}{max_row}"
    table_ref = f"{start_cell}:{end_cell}"

    try:
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        return True
    except Exception as e:
        print(f"    Warning: Could not create table {table_name}: {e}")
        return False


# =============================================================================
# Sheet Processing Functions
# =============================================================================


def _add_gpu_sheets(
    writer: pd.ExcelWriter,
    gpu_combined_path: Path,
    gpu_comparison_path: Path,
    verbose: bool,
) -> Tuple[List[str], List[str]]:
    """Add GPU timeline sheets.

    Returns (raw_sheets, comparison_sheets).
    """
    raw_sheets = []
    comparison_sheets = []

    if verbose:
        print("\nStep 1: Adding GPU Timeline sheets")

    # Read GPU combined (raw data)
    gpu_comb_xl = pd.ExcelFile(gpu_combined_path)
    for sheet_name in gpu_comb_xl.sheet_names:
        df = pd.read_excel(gpu_combined_path, sheet_name=sheet_name)
        new_name = GPU_SHEET_MAPPING.get(sheet_name, f"GPU_{sheet_name}_Raw")
        df.to_excel(writer, sheet_name=new_name, index=False)
        raw_sheets.append(new_name)
        if verbose:
            print(f"  Added {new_name} (will be hidden)")

    # Read GPU comparison
    gpu_comp_xl = pd.ExcelFile(gpu_comparison_path)
    for sheet_name in gpu_comp_xl.sheet_names:
        if "Comparison" in sheet_name:
            df = pd.read_excel(gpu_comparison_path, sheet_name=sheet_name)
            new_name = GPU_COMPARISON_MAPPING.get(sheet_name, f"GPU_{sheet_name}")
            df.to_excel(writer, sheet_name=new_name, index=False)
            comparison_sheets.append(new_name)
            if verbose:
                print(f"  Added {new_name}")

    return raw_sheets, comparison_sheets


def _add_collective_sheets(
    writer: pd.ExcelWriter,
    coll_combined_path: Path,
    coll_comparison_path: Path,
    verbose: bool,
) -> Tuple[List[str], List[str]]:
    """Add collective/NCCL sheets.

    Returns (raw_sheets, comparison_sheets).
    """
    raw_sheets = []
    comparison_sheets = []

    if verbose:
        print("\nStep 2: Adding Collective/NCCL sheets")

    # Read collective combined (raw data for hidden sheets)
    coll_comb_xl = pd.ExcelFile(coll_combined_path)
    for sheet_name in coll_comb_xl.sheet_names:
        if "summary" in sheet_name.lower():
            df = pd.read_excel(coll_combined_path, sheet_name=sheet_name)
            new_name = COLL_SHEET_MAPPING.get(sheet_name, f"NCCL_{sheet_name}_Raw")
            df.to_excel(writer, sheet_name=new_name, index=False)
            raw_sheets.append(new_name)
            if verbose:
                print(f"  Added {new_name} (will be hidden)")

    # Read collective comparison
    coll_comp_xl = pd.ExcelFile(coll_comparison_path)
    for sheet_name in coll_comp_xl.sheet_names:
        df = pd.read_excel(coll_comparison_path, sheet_name=sheet_name)

        # Determine appropriate naming
        if "nccl" in sheet_name.lower():
            if "_cmp" in sheet_name or "comparison" in sheet_name.lower():
                new_name = f"NCCL_{sheet_name.replace('nccl_', '').title().replace('_', '')}"
            else:
                new_name = f"NCCL_{sheet_name}"
        else:
            new_name = sheet_name

        df.to_excel(writer, sheet_name=new_name, index=False)

        if "_cmp" in sheet_name.lower() or "comparison" in sheet_name.lower():
            comparison_sheets.append(new_name)
            if verbose:
                print(f"  Added {new_name}")
        else:
            raw_sheets.append(new_name)
            if verbose:
                print(f"  Added {new_name} (will be hidden)")

    return raw_sheets, comparison_sheets


def _create_summary_dashboard(
    writer: pd.ExcelWriter,
    gpu_comparison_path: Path,
    coll_comparison_path: Path,
    baseline_label: str,
    test_label: str,
    verbose: bool,
) -> str:
    """Create Summary_Dashboard sheet with key metrics.

    Returns sheet name.
    """
    if verbose:
        print("\nStep 3: Creating Summary Dashboard")

    dashboard_data = {
        "Metric": [],
        baseline_label: [],
        test_label: [],
        "Improvement (%)": [],
        "Status": [],
    }

    # Add GPU metrics
    try:
        gpu_summary = pd.read_excel(gpu_comparison_path, sheet_name="Summary_Comparison")

        # Find the actual column names for time values
        time_cols = [
            col
            for col in gpu_summary.columns
            if "time_ms" in col and "diff" not in col and "percent" not in col
        ]

        if len(time_cols) >= 2:
            baseline_col = time_cols[0]
            test_col = time_cols[1]
        else:
            baseline_col = (
                "baseline_time_ms"
                if "baseline_time_ms" in gpu_summary.columns
                else time_cols[0] if time_cols else None
            )
            test_col = (
                "test_time_ms"
                if "test_time_ms" in gpu_summary.columns
                else time_cols[1] if len(time_cols) > 1 else None
            )

        if baseline_col and test_col:
            for _, row in gpu_summary.iterrows():
                metric_type = row["type"]
                dashboard_data["Metric"].append(f"GPU_{metric_type}")
                dashboard_data[baseline_label].append(round(row[baseline_col], 2))
                dashboard_data[test_label].append(round(row[test_col], 2))

                pct_val = row.get("percent_change", 0)
                dashboard_data["Improvement (%)"].append(round(pct_val, 2))

                if pct_val > 1:
                    status = "Better"
                elif pct_val < -1:
                    status = "Worse"
                else:
                    status = "Similar"
                dashboard_data["Status"].append(status)
    except Exception as e:
        if verbose:
            print(f"  Warning: Could not add GPU metrics to dashboard: {e}")

    # Add NCCL metrics
    try:
        # Try to read NCCL comparison sheets
        coll_xl = pd.ExcelFile(coll_comparison_path)
        nccl_cmp_sheets = [s for s in coll_xl.sheet_names if "_cmp" in s.lower()]

        for sheet_name in nccl_cmp_sheets:
            nccl_df = pd.read_excel(coll_comparison_path, sheet_name=sheet_name)

            # Find latency columns
            latency_cols = [
                col
                for col in nccl_df.columns
                if "comm_latency" in col.lower() and "percent_change" not in col.lower()
            ]

            if len(latency_cols) >= 2:
                base_col = latency_cols[0]
                test_col = latency_cols[1]
                pct_col = [
                    c
                    for c in nccl_df.columns
                    if "percent_change" in c.lower() and "latency" in c.lower()
                ]

                # Aggregate across all rows (mean)
                base_val = nccl_df[base_col].mean()
                test_val = nccl_df[test_col].mean()

                if pct_col:
                    pct_val = nccl_df[pct_col[0]].mean()
                else:
                    pct_val = (base_val - test_val) / base_val * 100 if base_val != 0 else 0

                # Create metric name from sheet name
                metric_name = sheet_name.replace("nccl_", "NCCL_").replace("_cmp", "_latency")

                dashboard_data["Metric"].append(metric_name)
                dashboard_data[baseline_label].append(round(base_val, 2))
                dashboard_data[test_label].append(round(test_val, 2))
                dashboard_data["Improvement (%)"].append(round(pct_val, 2))

                if pct_val > 1:
                    status = "Better"
                elif pct_val < -1:
                    status = "Worse"
                else:
                    status = "Similar"
                dashboard_data["Status"].append(status)

    except Exception as e:
        if verbose:
            print(f"  Warning: Could not add NCCL metrics to dashboard: {e}")

    dashboard_df = pd.DataFrame(dashboard_data)
    sheet_name = "Summary_Dashboard"
    dashboard_df.to_excel(writer, sheet_name=sheet_name, index=False)

    if verbose:
        print(f"  Added {sheet_name} ({len(dashboard_df)} metrics)")

    return sheet_name


def _apply_post_processing(
    output_path: Path,
    raw_sheets: List[str],
    comparison_sheets: List[str],
    verbose: bool,
) -> None:
    """Apply Excel formatting: hide sheets, add tables, color formatting."""

    if verbose:
        print("\nStep 4: Applying formatting")

    wb = load_workbook(output_path)

    # Hide raw data sheets
    for sheet_name in raw_sheets:
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_state = "hidden"
            if verbose:
                print(f"  Hidden: {sheet_name}")

    # Convert all sheets to tables and apply formatting
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip if sheet is empty
        if ws.max_row <= 1:
            continue

        # Create unique table name
        table_name = sanitize_table_name(sheet_name)

        if add_excel_table(ws, table_name):
            if verbose:
                print(f"  Converted to table: {sheet_name}")

        # Add conditional formatting for percent_change columns in comparison sheets
        if "Cmp" in sheet_name or "Comparison" in sheet_name or "Dashboard" in sheet_name:
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value and (
                    "percent_change" in str(cell_value).lower()
                    or "improvement" in str(cell_value).lower()
                ):
                    col_letter = get_column_letter(col_idx)
                    data_range = f"{col_letter}2:{col_letter}{ws.max_row}"

                    try:
                        ws.conditional_formatting.add(
                            data_range,
                            ColorScaleRule(
                                start_type="min",
                                start_color=RED,
                                mid_type="num",
                                mid_value=0,
                                mid_color=WHITE,
                                end_type="max",
                                end_color=GREEN,
                            ),
                        )
                        if verbose:
                            print(f"    Applied color scale to {sheet_name} column {cell_value}")
                    except Exception as e:
                        if verbose:
                            print(f"    Warning: Could not apply formatting to {cell_value}: {e}")

    # Move Summary Dashboard to first position
    if "Summary_Dashboard" in wb.sheetnames:
        dashboard_sheet = wb["Summary_Dashboard"]
        wb.move_sheet(dashboard_sheet, offset=-(len(wb.sheetnames) - 1))
        wb.active = 0  # Set dashboard as active sheet
        if verbose:
            print("\n  Moved Summary_Dashboard to first position")

    # Save workbook
    wb.save(output_path)


# =============================================================================
# Main Function
# =============================================================================


def create_final_excel_report(
    gpu_combined_path: Path,
    gpu_comparison_path: Path,
    coll_combined_path: Path,
    coll_comparison_path: Path,
    output_path: Path,
    baseline_label: str = "Baseline",
    test_label: str = "Test",
    verbose: bool = False,
) -> Dict[str, any]:
    """
    Create comprehensive final Excel report.

    Args:
        gpu_combined_path: Path to GPU combined file
        gpu_comparison_path: Path to GPU comparison file
        coll_combined_path: Path to collective combined file
        coll_comparison_path: Path to collective comparison file
        output_path: Output path for final report
        baseline_label: Label for baseline column
        test_label: Label for test column
        verbose: Print progress

    Returns:
        Dictionary with report metadata:
        - output_path: Path to created report
        - visible_sheets: List of visible sheet names
        - hidden_sheets: List of hidden sheet names
    """
    # Validate inputs
    for path, name in [
        (gpu_combined_path, "GPU combined"),
        (gpu_comparison_path, "GPU comparison"),
        (coll_combined_path, "Collective combined"),
        (coll_comparison_path, "Collective comparison"),
    ]:
        if not path.exists():
            raise FileNotFoundError(f"{name} file not found: {path}")

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Track sheets
    all_raw_sheets = []
    all_comparison_sheets = []

    # Create report
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Add GPU sheets
        gpu_raw, gpu_cmp = _add_gpu_sheets(writer, gpu_combined_path, gpu_comparison_path, verbose)
        all_raw_sheets.extend(gpu_raw)
        all_comparison_sheets.extend(gpu_cmp)

        # Add collective sheets
        coll_raw, coll_cmp = _add_collective_sheets(
            writer, coll_combined_path, coll_comparison_path, verbose
        )
        all_raw_sheets.extend(coll_raw)
        all_comparison_sheets.extend(coll_cmp)

        # Create summary dashboard
        dashboard_sheet = _create_summary_dashboard(
            writer,
            gpu_comparison_path,
            coll_comparison_path,
            baseline_label,
            test_label,
            verbose,
        )

    # Apply post-processing (hide sheets, add tables, formatting)
    _apply_post_processing(output_path, all_raw_sheets, all_comparison_sheets, verbose)

    return {
        "output_path": output_path,
        "visible_sheets": [dashboard_sheet] + all_comparison_sheets,
        "hidden_sheets": all_raw_sheets,
    }
