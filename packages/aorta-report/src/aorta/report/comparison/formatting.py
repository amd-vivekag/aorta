"""Shared Excel formatting utilities for comparison reports."""

from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule


# Color constants for conditional formatting
RED = "F8696B"
WHITE = "FFFFFF"
GREEN = "63BE7B"


def get_column_letter(col_idx: int) -> str:
    """
    Convert 1-based column index to Excel column letter.

    Args:
        col_idx: 1-based column index

    Returns:
        Excel column letter (A, B, ..., Z, AA, AB, ...)

    Examples:
        >>> get_column_letter(1)
        'A'
        >>> get_column_letter(26)
        'Z'
        >>> get_column_letter(27)
        'AA'
        >>> get_column_letter(28)
        'AB'
    """
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def create_color_scale_rule() -> ColorScaleRule:
    """
    Create standard red-white-green color scale rule.

    Red (min/negative) → White (0) → Green (max/positive)

    Returns:
        ColorScaleRule configured for percent_change columns
    """
    return ColorScaleRule(
        start_type="min",
        start_color=RED,
        mid_type="num",
        mid_value=0,
        mid_color=WHITE,
        end_type="max",
        end_color=GREEN,
    )


def apply_color_scale_to_column(
    worksheet,
    col_idx: int,
    num_rows: int,
) -> None:
    """
    Apply color scale formatting to a specific column.

    Args:
        worksheet: openpyxl worksheet
        col_idx: 1-based column index
        num_rows: Number of data rows (excluding header)
    """
    col_letter = get_column_letter(col_idx)
    # Data starts at row 2 (row 1 is header)
    data_range = f"{col_letter}2:{col_letter}{num_rows + 1}"

    worksheet.conditional_formatting.add(data_range, create_color_scale_rule())


def save_with_formatting(
    data: Dict[str, pd.DataFrame],
    output_path: Path,
    format_columns: Dict[str, List[str]],
    verbose: bool = False,
) -> Path:
    """
    Save DataFrames to Excel with conditional formatting.

    Args:
        data: Dict[sheet_name, DataFrame]
        output_path: Output file path
        format_columns: Dict[sheet_name, list of column names to format]
        verbose: Print progress

    Returns:
        Path to saved file

    Example:
        format_columns = {
            "Comparison_By_Rank": ["percent_change"],
            "Summary_Comparison": ["percent_change"],
            "nccl_implicit_sync_cmp": [
                "percent_change_comm_latency_mean",
                "percent_change_algo bw (GB/s)_mean",
            ],
        }
    """
    # Ensure output directory exists
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write all sheets
        for sheet_name, df in data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply formatting
        for sheet_name, columns_to_format in format_columns.items():
            if sheet_name not in data:
                continue

            df = data[sheet_name]
            worksheet = writer.sheets[sheet_name]
            num_rows = len(df)

            for col_name in columns_to_format:
                if col_name not in df.columns:
                    continue

                # Find column index (1-based)
                col_idx = df.columns.get_loc(col_name) + 1

                apply_color_scale_to_column(worksheet, col_idx, num_rows)

                if verbose:
                    print(f"    Formatted {sheet_name}.{col_name}")

    if verbose:
        print(f"\nSaved: {output_path}")

    return output_path
